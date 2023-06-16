# -*- coding: utf-8 -*-
"""
Created on Fri May  5 15:32:01 2023

@author: raulgranja
"""
from pathlib import Path
from openpyxl import load_workbook
import chromedriver_autoinstaller
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from bs4 import BeautifulSoup

# Loading Path
current_path = Path.cwd()
input_sheet = current_path / Path('input.xlsx')

# Loading input
wb_i = load_workbook(filename = input_sheet)
ws_i = wb_i.active
upns = ws_i['A']

# Creating output
output_sheet = current_path / Path('output.xlsx')
wb_o = load_workbook(filename = output_sheet)
ws_o = wb_o.active

# Verify Chrome Driver
chromedriver_autoinstaller.install()

# Comment code below to see Chrome popup
options = Options()
options.add_argument("--window-size=1920,1080")
options.add_argument("--headless")

driver = webdriver.Chrome(options=options)

# Acess Website
site = "Lorem"
driver.get(site)
print("Chrome Initialized")

# Initializing list of rows
data = list()

# Loop through UPNs
for upn in upns:
    upn = str(upn.value)

    if upn == None:
        break

    else:
        # Search UPN
        material = driver.find_element_by_id("Material")
        material.send_keys(upn)
        search = driver.find_element_by_class_name("SmallButtonFixed")
        search.click()
        page_source = driver.page_source

        # Scraping table
        soup = BeautifulSoup(page_source, 'lxml')
        table = soup.find_all('table')[7]

        for list_index, i in enumerate(table.tbody.find_all('tr')):
            if list_index == 0:
                continue
            # Find all data for each row 
            row = [j.text.strip() for j in i.find_all('td')]
            ws_o.append(row)
            data.append(row)

        driver.find_element_by_id("Material").clear()

# Creating DF
columns = ['Ipsum', 'Dolor', 'Sit', 'Amet', 'Consectetur',
           'Adipiscing', 'Elit']

wb_o.save(filename='output.xlsx')
